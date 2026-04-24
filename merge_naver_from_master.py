# -*- coding: utf-8 -*-
"""master xlsx에서 네이버스토어 제품만 추출해 products.json에 병합.
- 기존 products.json 은 백업 (products.backup.json)
- og:image 도 바로 채움 (smartstore.naver.com 대상)
- 중복 방지: product + officialLink 키로 체크
"""
import json
import re
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
REPO_ROOT = HERE.parent
MASTER = REPO_ROOT / "data" / "supergredients_master.xlsx"
JSON_PATH = HERE / "products.json"
BACKUP = HERE / "products.backup.json"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Accept-Language": "ko,en;q=0.9",
}
OG_PATTERNS = [
    re.compile(r'<meta[^>]*property=["\']og:image["\'][^>]*content=["\']([^"\']+)["\']', re.I),
    re.compile(r'<meta[^>]*content=["\']([^"\']+)["\'][^>]*property=["\']og:image["\']', re.I),
    re.compile(r'<meta[^>]*name=["\']twitter:image["\'][^>]*content=["\']([^"\']+)["\']', re.I),
]


def fetch_og(url: str, timeout: int = 10) -> str:
    if not url:
        return ""
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read(500_000)
        html = raw.decode("utf-8", errors="ignore")
        for pat in OG_PATTERNS:
            m = pat.search(html)
            if m:
                img = m.group(1).strip()
                if img.startswith("//"):
                    img = "https:" + img
                return img
    except Exception:
        pass
    return ""


def main():
    with JSON_PATH.open("r", encoding="utf-8") as f:
        products = json.load(f)

    # 백업
    with BACKUP.open("w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    existing_keys = set()
    for p in products:
        existing_keys.add((p.get("product", "").strip(), p.get("officialLink", "").strip()))

    max_no = 0
    for p in products:
        try:
            n = int(str(p.get("no", "0")).strip() or 0)
            if n > max_no:
                max_no = n
        except ValueError:
            pass

    wb = openpyxl.load_workbook(MASTER, data_only=True)
    ws = wb["supergredients"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def g(row, key, default=""):
        i = idx.get(key)
        if i is None:
            return default
        v = row[i]
        return "" if v is None else str(v).strip()

    added = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        plat = g(row, "플랫폼")
        if "네이버" not in plat:
            continue
        product = g(row, "상품")
        if not product:
            continue
        official = g(row, "공식링크")
        affiliate = g(row, "어필리에이트 링크")
        key = (product, official)
        if key in existing_keys:
            continue

        max_no += 1
        # 플랫폼 이름 정리 ("7.  네이버스토어" → "5. 네이버스토어") — 기존 1~4와 번호 이어지게
        plat_clean = "5. 네이버스토어"

        rec = {
            "no": str(max_no),
            "platform": plat_clean,
            "category": g(row, "카테고리"),
            "product": product,
            "ingredients": g(row, "원재료"),
            "clean": g(row, "CLEAN"),
            "cleanReason": g(row, "첨가물 0개 (완전 클린)"),
            "note": g(row, "참고노트"),
            "fullName": g(row, "상품명") or product,
            "affiliateLink": affiliate,
            "officialLink": official,
            "volume": g(row, "용량"),
            "price": g(row, "판매가 (원)"),
            "producer": g(row, "생산자"),
            "origin": g(row, "원산지"),
            "imageUrl": "",
        }
        # og:image
        img = fetch_og(official) or fetch_og(affiliate)
        rec["imageUrl"] = img
        products.append(rec)
        added.append(rec)
        existing_keys.add(key)
        print(f"  + [{rec['category']}] {product} — image: {'ok' if img else 'no'}")

    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    print(f"\n[+] added {len(added)} 네이버 products. total now: {len(products)}")


if __name__ == "__main__":
    main()
